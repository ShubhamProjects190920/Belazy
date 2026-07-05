"""
Excel Service — Import and export projects via .xlsx files.
"""
from datetime import date
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = [
    "Name *",
    "Client Name",
    "Location",
    "Status",
    "Priority",
    "Start Date (YYYY-MM-DD)",
    "End Date (YYYY-MM-DD)",
    "Budget (INR)",
    "Description",
]

VALID_STATUSES  = {"planning", "active", "on_hold", "completed", "canceled"}
VALID_PRIORITIES = {"low", "medium", "high", "critical"}

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1E3A5F")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_EXAMPLE_FONT = Font(color="888888", italic=True)


def _safe_str(val: Any) -> str | None:
    """Return stripped string or None for blank/None values."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() != "none" else None


def _style_header_row(ws) -> None:
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 24


def export_projects(projects: list) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"
    ws.freeze_panes = "A2"

    _style_header_row(ws)

    for row_num, project in enumerate(projects, 2):
        ws.cell(row=row_num, column=1, value=project.name)
        ws.cell(row=row_num, column=2, value=project.client_name or "")
        ws.cell(row=row_num, column=3, value=project.location or "")
        ws.cell(row=row_num, column=4, value=project.status.value)
        ws.cell(row=row_num, column=5, value=project.priority.value)
        ws.cell(row=row_num, column=6, value=str(project.start_date) if project.start_date else "")
        ws.cell(row=row_num, column=7, value=str(project.end_date) if project.end_date else "")
        ws.cell(row=row_num, column=8, value=float(project.budget) if project.budget is not None else "")
        ws.cell(row=row_num, column=9, value=project.description or "")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_template() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"
    ws.freeze_panes = "A2"

    _style_header_row(ws)

    # Example row so user knows the expected format
    example = [
        "Highway Bridge Project",
        "NHAI",
        "Mumbai, Maharashtra",
        "planning",
        "high",
        "2026-08-01",
        "2027-06-30",
        "5000000",
        "Construction of 4-lane bridge over river",
    ]
    for col, value in enumerate(example, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.font = _EXAMPLE_FONT

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_import(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """
    Parse an uploaded Excel file into a list of project dicts and a list of error messages.
    Row 1 is the header and is always skipped. Row 2 onwards is data.
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    projects: list[dict] = []
    errors: list[str] = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        # Skip completely blank rows
        if not any(cell for cell in row):
            continue

        name = _safe_str(row[0])
        if not name:
            errors.append(f"Row {row_num}: 'Name' is required — row skipped.")
            continue

        # Status — default to planning if invalid
        raw_status = _safe_str(row[3])
        status = raw_status if raw_status in VALID_STATUSES else "planning"

        # Priority — default to medium if invalid
        raw_priority = _safe_str(row[4])
        priority = raw_priority if raw_priority in VALID_PRIORITIES else "medium"

        # Dates
        start_date = None
        end_date = None
        raw_start = _safe_str(row[5])
        raw_end = _safe_str(row[6])
        if raw_start:
            try:
                start_date = date.fromisoformat(raw_start)
            except ValueError:
                errors.append(f"Row {row_num}: Invalid start date '{raw_start}' — ignored.")
        if raw_end:
            try:
                end_date = date.fromisoformat(raw_end)
            except ValueError:
                errors.append(f"Row {row_num}: Invalid end date '{raw_end}' — ignored.")

        # Budget
        budget = None
        raw_budget = row[7]
        if raw_budget is not None and str(raw_budget).strip() not in ("", "None"):
            try:
                budget = float(raw_budget)
            except (ValueError, TypeError):
                errors.append(f"Row {row_num}: Invalid budget value — ignored.")

        projects.append({
            "name":        name,
            "client_name": _safe_str(row[1]),
            "location":    _safe_str(row[2]),
            "status":      status,
            "priority":    priority,
            "start_date":  start_date,
            "end_date":    end_date,
            "budget":      budget,
            "description": _safe_str(row[8]),
        })

    return projects, errors
